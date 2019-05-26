#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl as op
from openpyxl import Workbook as wob

from src.bo_dau.xuly_bodau import no_accent as na


def out():
    wb = op.load_workbook('D:/TK 13.3/Doantotnghiep/Code/demo/file/data/data.xlsx')

    ws = wob.active
    x = wb.get_sheet_names()

    my_wb = op.Workbook()
    my_sheet = my_wb.active

    sheet = wb.get_sheet_by_name(x[0])

    for x in range(1, sheet.max_row):
        kq = na(sheet.cell(row=x, column=1).value)
        my_sheet.cell(row=x, column=1).value = kq
        my_sheet.cell(row=x, column=2).value = sheet.cell(row=x, column=2).value

    my_wb.save('D:/TK 13.3/Doantotnghiep/Code/demo/file/data/data_no_acc.xlsx')
