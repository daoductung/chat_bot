#!/usr/bin/env python
# -*- coding: utf-8 -*-
import glob

import pandas as pd
from src.model.svm_model import SVMModel
from sklearn.externals import joblib
from src.bo_dau import demo_writeexcel
import openpyxl as op


class TextClassificationPredict(object):
    def __init__(self):
        self.test = None

    def get_train_data(self):
        demo_writeexcel.out()
        #  train data
        train_data = []
        file = glob.glob("D:/TK 13.3/Doantotnghiep/Code/demo/file/data/*.xlsx")
        for f in file:
            wb = op.load_workbook(f)
            x = wb.get_sheet_names()
            sheet = wb.get_sheet_by_name(x[0])
            for x in range(1, sheet.max_row + 2):
                if sheet.cell(row=x, column=1).value is not None:
                    train_data.append(
                        {"body": sheet.cell(row=x, column=1).value, "label": sheet.cell(row=x, column=2).value})
        df_train = pd.DataFrame(train_data)
        print(df_train)
        model = SVMModel()
        clf = model.clf.fit(df_train["body"], df_train['label'])
        joblib.dump(clf, 'D:/TK 13.3/Doantotnghiep/Code/demo/save/data.pkl')
        print("Xong")


if __name__ == '__main__':
    tcp = TextClassificationPredict()
    tcp.get_train_data()
