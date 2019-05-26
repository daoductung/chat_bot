import re

'''import glob
import openpyxl as op

sai = []
file = "D:/TK 13.3/Doantotnghiep/Code/demo/reply/reply.xlsx"

wb = op.load_workbook(file)
print(wb)
x = wb.get_sheet_names()
print(x)
sheet = wb.get_sheet_by_name('Sai')
for x in range(1, sheet.max_row + 1):
    if sheet.cell(row=x, column=1).value is not None:
        sai.append(sheet.cell(row=x, column=1).value)
print(sai)'''
'''message = {}
attachment = {}
type = {}
payload = {}
buttons = []
data = {}

message['message'] = attachment
attachment['attachment'] = type
type['type'] = 'template'
type['payload'] = payload
payload['template_type'] = 'button'
payload['text'] = 'Hãy chọn nhu cầu phù hợp với bạn!'
payload['buttons'] = buttons
data['type'] = 'web_url'
data['url'] = 'abc.com'
data['title'] = 'abc'
buttons.append(data.copy())

print(message)



def return_message(template_type=None, content=None, template=None, dict =dict):
    message = {}
    attachment = {}
    type = {}
    payload = {}
    buttons = []
    data = {}

    message['message'] = data
    if template_type == 'text':
        data[template_type] = content
    elif template_type == 'button':
        message['message'] = attachment
        attachment['attachment'] = type
        type['type'] = template
        type['payload'] = payload
        payload['template_type'] = template_type
        payload['text'] = content
        payload['buttons'] = buttons
        for x in dict:
            print(x)
            buttons.append(x)

    return message

data = []
user_data = {'ID': 2048, 'Name': 'john', 'Father Name':'David'}
data.append(user_data)
new_user_data = {'ID': 2049, 'Name':'bob', 'Father Name':'McBobsalot'}
data.append(new_user_data)
print(data)
print(return_message(template_type='button', content='Hãy chọn nhu cầu phù hợp!', template='template', dict=data))

def add_button(*args):
    data=[]
    for x in range(0, len(args)):
        data.append(args[x])
    return data


button1 = {'type': 'text', 'title': 'ten'}
button2 = {'type': 'text', 'title': 'tuoi'}
button3 = {'type': 'text', 'title': 'hi'}

print(add_button(button2, button1,button3))

intent = []

thong_tin_truong = []

lich_su_pt = []
vi_tri = []
thong_tin_noi_bat = []
co_so_vat_chat = []
hoc_phi = []
co_hoi_viec_lam = []
hoat_dong = []

thong_tin_tuyensinh = []
diem_chuan = []
diem_chuan_nam = []
phuong_thuc_tuyen_sinh = []
ho_so = []
cach_nop_ho_so = []
chi_tieu_tuyen_sinh = []

thong_tin_nganhhoc = []
ten_nganh = []
gioi_thieu_nganh = []
chi_tieu_tuyen_sinh_khoa = []
hoat_dong_khoa = []
co_hoi_viec_lam_khoa = []

gt_thong_tin_chung = []
gt_nhiem_vu_chuyen_mon = []
gt_huong_nghien_cuu = []
gt_doi_ngu_gv = []
'''
'''
import openpyxl as op

cau_tra_loi = []
file = "D:/TK 13.3/Doantotnghiep/Code/demo/reply/reply.xlsx"
wb = op.load_workbook(file)
x = wb.get_sheet_names()
for i in x:
    print(i)
    if i != '':
        sheet = wb.get_sheet_by_name(i)
        print(1)
        for x in range(1, sheet.max_row + 1):
            if sheet.cell(row=x, column=2).value is not None:
                print(2)
'''

'''def test(*args):
    print(len(args))
    print(args)
    for i in args:
        if i != '':
            print(i)

test()'''
'''
import re
re_chi_tieu = r'(?<=)((chỉ\s+tiêu)|(bao nhiêu sinh viên)|(tuyển\s+chọn))'
t = re.search(re_chi_tieu, 'chỉ tiêu tuyển sinh đại học')
if t is None:
    print('h')
'''

'''
def convert(list):
    # Converting integer list to string list
    # and joining the list using join()
    res = str(''.join(map(str, list)))

    return res


# Driver code

print(convert(list))

from handling import handling as h
re_nam = r'\d*\s*([nN][ăaĂ][mM])(\s*([nN][aA][yY]|[hH][iI][êÊệỆ][nN]\s*[tT][ạẠ][iI]|[tT][rR][ưƯứỨ][ơƠớỚ][cC]|[sS][aA][uU]|[nN][gG][oO][áÁ][iI]))(?=$|\s|\,|\.)|(([nN][ăaĂ][mM])*\s*(20\d+))'
chuoi = "năm sau"
so = r'20\d+'
time =r'trước|nay|sau'

print(h.regex_search(re_nam, chuoi))
print(h.regex_search(time, chuoi))
'''
import time

def stopwatch(seconds):
    start = time.time()
    print(time.clock())
    elapsed = 0
    while elapsed == seconds:
        elapsed = time.time() - start
        print("loop cycle time: %f, seconds count: %02d" % (time.clock(), elapsed))

stopwatch(20)
