#!/usr/bin/env python
# -*- coding: utf-8 -*-

from random import choice as ch
from intent.demo import identify_intention
import openpyxl as op
import re
import requests
import bs4
from handling.handling_file import wirte_file_excel


def crawl_point(nam):
    js = {}
    url = 'https://diemthi.tuyensinh247.com/diem-chuan/dai-hoc-su-pham-ky-thuat-hung-yen-SKH.html?y=' + str(nam)
    response = requests.get(url)
    html = response.content

    soup = bs4.BeautifulSoup(html, "lxml")
    for link in soup.find_all('table'):
        for tr in link.find_all('tr')[1:2]:
            if tr.get_text().strip() != '':
                js['DH'] = tr.find_all('td')[4:6][0].get_text()
                js['THPT'] = tr.find_all('td')[4:6][1].get_text()

    return js


def regex_search(regex, chuoi):
    t = re.search(regex, chuoi)
    if t is not None:
        group = t.group(0)
        start = t.start(0)
        end = t.end(0)
    else:
        group = None
    return group


def give_answer(*args, kichban=str):
    cau_tra_loi = []
    file = "D:/TK 13.3/Doantotnghiep/Code/demo/file/reply/reply.xlsx"
    wb = op.load_workbook(file)
    x = wb.get_sheet_names()
    for i in x:
        if i == kichban:
            sheet = wb.get_sheet_by_name(kichban)
            if sheet:
                for x in range(1, sheet.max_row + 1):
                    if sheet.cell(row=x, column=2).value is not None:
                        for h in args:
                            if h is not None:
                                if sheet.cell(row=x, column=2).value == h:
                                    cau_tra_loi.append(sheet.cell(row=x, column=1).value)

                    elif sheet.cell(row=x, column=2).value is None:
                        if sheet.cell(row=x, column=1).value is not None:
                            cau_tra_loi.append(sheet.cell(row=x, column=1).value)
    return cau_tra_loi


def checknotnull(*args):
    for x in args:
        if x is not None:
            return x


def handling_year(chuoi):
    if chuoi == 'nay' or chuoi == '2019':
        pass


def return_reply(intent=[], sentence_activated=[]):
    print(intent)
    print(sentence_activated)
    message = {}
    body = sentence_activated[-1]
    kichban = intent[-1]
    print('câu hỏi kích hoạt gần nhất: {0}, kịch bản: {1}'.format(body, kichban))
    give_diem = []
    '''
        Phần regex tuyển sinh
    '''
    re_chi_tieu = r'(?<=)((chỉ\s+tiêu)|(bao nhiêu sinh viên)|(tuyển\s+chọn)|(tuyển\s+sinh\s+bao\s+nhiêu))'
    re_nhu_cau_diem = r'(?<=)(điểm)'
    re_diem = r'((điểm|diem)\s*\d{2})|(\d{2}\s*(điểm|diem)|(\d{2}\s*(đ|d)))'
    re_nam = r'\d*\s*([nN][ăaĂ][mM])(\s*([nN][aA][yY]|[hH][iI][êÊệỆ][nN]\s*[tT][ạẠ][iI]|[tT][rR][ưƯứỨ][ơƠớỚ][cC]|[sS][aA][uU]|[nN][gG][oO][áÁ][iI]))(?=$|\s|\,|\.)|(([nN][ăaĂ][mM])*\s*(20\d+))|(những năm gần đây)'
    re_ho_so = r'(?<=)(hồ\s+sơ)'
    re_cach_nop_ho = r'(?<=)((nộp\s+hồ\s+sơ)|(nộp\s+vào\s+thời\s+gian)|(nộp))'
    re_phuong_thuc_tuyen_sinh = r'(?<=)((hình\s+thức)|(phương\s+thức)|(học\s+bạ))'
    re_dieu_kien_xet_tuyen = r'(?<=)((điều kiện xét tuyển)|(dkxt))'
    re_ma = r'(mã)'
    '''
        Phần regex thông tin về trường
    '''

    re_lich_su_phat_trien = r'(?<=)((thành lập)|(từ năm)|(lịch sử))'
    re_vi_tri = r'(?<=)((nằm ở đâu)|(nằm chỗ nào)|(gần chỗ)|(vị trí))'
    re_thong_tin = r'(?<=)((thông\s+tin)|(giới\s+thiệu\s+qua\s+về\s+trường)|(nổi\s+bật)|(thành\s+tích))'
    re_co_so_vat_chat = r'(?<=)((cơ\s+sở\s+vật\s+chất)|(đồ\s+dùng)|(cơ\s+sở)|(sân)|(thể\s+thao))'
    re_hoc_phi = r'(?<=)(học\s+phí)'
    re_co_hoi_viec_lam = r'(?<=)((việc\s+làm)|(cơ\s+hội\s+việc\s+làm))'
    re_hoat_dong_doan_the = r'(?<=)((hoạt\s+động)|đoàn)'

    '''
        Phần regex thông tin ngành học
    '''
    re_he_thong_nhung = r'(?<=)(nhúng|hệ thống nhúng|làm mạch|điều khiển|iot|smart\s+home|công nghệ máy tính|kỹ thuật máy tính)'
    re_lap_trinh_web = r'(?<=)(công nghệ phần mềm|phần mềm|web|website)|(di động|mobile|android|ios|ứng dụng|app)|(kiểm thử phần mềm|tester|test)'
    re_lap_trinh_mang = r'(?<=)(mạng)'
    re_lap_trinh_mobile = r'(?<=)(di động|mobile|android|ios|ứng dụng|app)|(kiểm thử phần mềm|tester|test)'
    re_kiem_thu = r'(?<=)(kiểm thử phần mềm|tester|test)'
    re_hung_yen_aptech = r'(?<=)(hưng yên aptech)'
    re_khoi_xet_tuyen = r'(?<=)(khối)'
    re_co_so_dao_tao = r'(?<=)(đào tạo ở đâu|học ở đâu|học chỗ nào|học ở chỗ nào)'
    re_nhiem_vu_chuyen_mon = r'(?<=)(nhiệm vụ)'
    re_huong_nghien_cuu = r'(?<=)(hướng nghiên cứu)'
    re_doi_ngu_can_bo = r'(?<=)(đội ngũ cán bộ|giáo viên|giảng viên)'
    re_kien_thuc_ky_nang = r'(?<=)(kiến thức|kỹ năng)'
    re_vi_tri_lam_viec = r'(?<=)(làm ở vị trí nào|vị trí làm việc)'
    re_thong_tin_chung = r'(?<=)(thông tin chung|giới thiệu|thông tin)'
    '''
    '''
    if kichban == 'thong_tin_ve_truong':
        print(1)
        lich_su_pt = regex_search(re_lich_su_phat_trien, body)
        vi_tri = regex_search(re_vi_tri, body)
        thong_tin_noi_bat = regex_search(re_thong_tin, body)
        co_so_vat_chat = regex_search(re_co_so_vat_chat, body)
        hoc_phi = regex_search(re_hoc_phi, body)
        co_hoi_viec_lam = regex_search(re_co_hoi_viec_lam, body)
        hoat_dong = regex_search(re_hoat_dong_doan_the, body)
        if lich_su_pt is not None:
            lich_su_pt = 'lich_su_phat_trien'
        if vi_tri is not None:
            vi_tri = 'vi_tri'
        if thong_tin_noi_bat is not None:
            thong_tin_noi_bat = 'giai_thuong'
        if co_so_vat_chat is not None:
            co_so_vat_chat = 'co_so_vat_chat'
        if hoc_phi is not None:
            hoc_phi = 'hoc_phi'
        if co_hoi_viec_lam is not None:
            co_hoi_viec_lam = 'co_hoi_viec_lam'
        if hoat_dong is not None:
            hoat_dong = 'hoat_dong_truong'

        if lich_su_pt is None and vi_tri is None and thong_tin_noi_bat is None and co_so_vat_chat is None and hoc_phi is None and co_hoi_viec_lam is None and hoat_dong is None:
            button1 = {'type': 'text', 'title': 'Lịch sử phát triển'}
            button2 = {'type': 'text', 'title': 'Vị trí của trường'}
            button3 = {'type': 'text', 'title': 'Thành tích của trường'}
            button4 = {'type': 'text', 'title': 'Cơ sở vật chất'}
            button5 = {'type': 'text', 'title': 'Học phí của trường'}
            button6 = {'type': 'text', 'title': 'Việc làm sau khi ra trường'}
            button7 = {'type': 'text', 'title': 'Hoạt động đoàn thể của trường'}
            dict = add_button(button1, button2, button3, button4, button5, button6, button7)
            message = return_message(template_type='button',
                                     content=['Hãy chọn một trong những thông tin sau để chúng tôi tư vấn:'],
                                     template='template', dict=dict)

        else:
            result = give_answer(lich_su_pt, vi_tri, thong_tin_noi_bat, co_so_vat_chat, hoc_phi, co_hoi_viec_lam,
                                 hoat_dong,
                                 kichban='thong_tin_ve_truong')

            message = return_message(template_type='text', content=[ch(result)])


    elif kichban == 'tuyen_sinh':
        chi_tieu_tuyen_sinh = regex_search(re_chi_tieu, body)
        nhu_cau_diem = regex_search(re_nhu_cau_diem, body)
        ho_so = regex_search(re_ho_so, body)
        cach_nop_ho_so = regex_search(re_cach_nop_ho, body)
        phuong_thuc_tuyen_sinh = regex_search(re_phuong_thuc_tuyen_sinh, body)
        dieu_kien_xet_tuyen = regex_search(re_dieu_kien_xet_tuyen, body)
        if chi_tieu_tuyen_sinh is not None:
            chi_tieu_tuyen_sinh = 'chi_tieu_tuyen_sinh'
        if nhu_cau_diem is not None:
            nhu_cau_diem = 'diem_tuyen_sinh'
            nam = regex_search(re_nam, body)

        if ho_so is not None:
            ho_so = 'ho_so_xet_tuyen'
        if cach_nop_ho_so is not None:
            cach_nop_ho_so = 'cach_nop_ho_so'
        if phuong_thuc_tuyen_sinh is not None:
            phuong_thuc_tuyen_sinh = 'phuong_thuc_tuyen_sinh'
        if dieu_kien_xet_tuyen is not None:
            dieu_kien_xet_tuyen = 'dieu_kien_xet_tuyen'

        if chi_tieu_tuyen_sinh is None and nhu_cau_diem is None and ho_so is None and cach_nop_ho_so is None and phuong_thuc_tuyen_sinh is None and dieu_kien_xet_tuyen is None:
            button1 = {'type': 'text', 'title': 'Chỉ tiêu tuyển sinh'}
            button2 = {'type': 'text', 'title': 'Điểm tuyển sinh'}
            button3 = {'type': 'text', 'title': 'Hồ sơ tuyển sinh'}
            button4 = {'type': 'text', 'title': 'Cách nộp hồ sơ'}
            button5 = {'type': 'text', 'title': 'Phương thức xét tuyển'}
            button6 = {'type': 'text', 'title': 'Điều kiện xét tuyển'}
            dict = add_button(button1, button2, button3, button4, button5, button6)
            message = return_message(template_type='button',
                                     content=['Hãy chọn một trong những thông tin sau để chúng tôi tư vấn:'],
                                     template='template', dict=dict)
        else:
            if nhu_cau_diem is not None:

                if nam is not None:
                    nam=nam.strip()
                    if nam == 'năm nay' or nam == 'năm 2019' or nam == 'năm hiện tại':
                        diem_thi = crawl_point(2019)
                    elif nam == 'năm trước' or nam == 'năm ngoái' or nam == 'năm 2018':
                        diem_thi = crawl_point(2018)
                    elif nam == 'năm 2017':
                        diem_thi = crawl_point(2017)
                    elif nam == 'những năm gần đây':
                        for i in [2016, 2017, 2018]:
                            give_diem.append(crawl_point(i))
                    elif nam == 'năm sau':
                        diem_thi = crawl_point(2020)

                    give_diem.append(diem_thi)
                    print(give_diem)
                    diem = regex_search(re_diem, body)
                    if diem is not None:
                        if len(give_diem[-1]) > 0:
                            if int(regex_search(r'\d{2}', diem)) >= int(give_diem[-1]['DH']):
                                text = 'Với mức điểm {0}, bạn đủ điểm đỗ khoa CNTT năm {1}'.format(diem, nam)
                            else:
                                text = 'Với mức điểm {0}, bạn không đủ điểm đỗ khoa CNTT năm {1}'.format(diem, nam)
                        else:
                            text = 'Hiện tại chúng tôi không thể tìm thấy điểm của {0}. Bạn có thể truy câp <a target="_blank" href="https://diemthi.tuyensinh247.com/diem-chuan/dai-hoc-su-pham-ky-thuat-hung-yen-SKH.html"><b>tại đây</b></a>'.format(
                                nam)

                        message = return_message(template_type='text', content=[text])
                    else:
                        if len(give_diem[-1]) > 0:
                            text = 'Điểm xét tuyển đại học {0} là {1}, {2}'.format(nam, give_diem[-1]['DH'], give_diem[-1]['THPT'])
                        else:
                            text = 'Hiện tại chúng tôi không thể tìm thấy điểm của năm {0}. Bạn có thể truy câp <a target="_blank" href="https://diemthi.tuyensinh247.com/diem-chuan/dai-hoc-su-pham-ky-thuat-hung-yen-SKH.html">tại đây</a> để xem điểm'.format(
                                nam)

                        message = return_message(template_type='text', content=[text])
                else:
                    message = return_message(template_type='text',
                                             content=['Bạn muốn biết điểm xét tuyển năm nào?'])
            else:
                result = give_answer(chi_tieu_tuyen_sinh, ho_so, cach_nop_ho_so, phuong_thuc_tuyen_sinh,
                                     dieu_kien_xet_tuyen, kichban='tuyen_sinh')

                message = return_message(template_type='text', content=result)
    elif kichban == 'khoa_cntt':
        chuyen_nganh = []

        ky_thuat_may_tinh = regex_search(re_he_thong_nhung, body)
        cong_nghe_phan_mem = regex_search(re_lap_trinh_web, body)
        mang = regex_search(re_lap_trinh_mang, body)
        hung_yen_aptech = regex_search(re_hung_yen_aptech, body)

        chi_tieu_cntt = regex_search(re_chi_tieu, body)
        khoi_xet_tuyen = regex_search(re_khoi_xet_tuyen, body)
        hoat_dong_doan = regex_search(re_hoat_dong_doan_the, body)
        co_so_dao_tao = regex_search(re_co_so_dao_tao, body)

        thong_tin_chung = regex_search(re_thong_tin_chung, body)
        nhiem_vu_chuyen_mon = regex_search(re_nhiem_vu_chuyen_mon, body)
        huong_nghien_cuu = regex_search(re_huong_nghien_cuu, body)
        doi_ngu_can_bo = regex_search(re_doi_ngu_can_bo, body)
        kien_thuc_ky_nang = regex_search(re_kien_thuc_ky_nang, body)
        vi_tri_lam_viec = regex_search(re_vi_tri_lam_viec, body)

        if hoat_dong_doan is not None:
            hoat_dong_doan = 'hoat_dong_ngoai_khoa'
        if chi_tieu_cntt is not None:
            chi_tieu_cntt = 'chi_tieu_cntt'
        if khoi_xet_tuyen is not None:
            khoi_xet_tuyen = 'to_hop_mon_xet_tuyen'
        if co_so_dao_tao is not None:
            co_so_dao_tao = 'co_so_dao_tao'
        if ky_thuat_may_tinh is not None:
            ky_thuat_may_tinh = 'ky_thuat_may_tinh'
        if cong_nghe_phan_mem is not None:
            cong_nghe_phan_mem = 'cong_nghe_phan_mem'
        if mang is not None:
            mang = 'mang_may_tinh_va_truyen_thong'
        if hung_yen_aptech is not None:
            hung_yen_aptech = 'Hung_Yen_Aptech'
        if nhiem_vu_chuyen_mon is not None:
            nhiem_vu_chuyen_mon = 'nhiem_vu_chuyen_mon'
        if huong_nghien_cuu is not None:
            huong_nghien_cuu = 'cac_huong_nghien_cuu_chinh'
        if doi_ngu_can_bo is not None:
            doi_ngu_can_bo = 'doi_ngu_can_bo'
        if kien_thuc_ky_nang is not None:
            kien_thuc_ky_nang = 'kien_thuc_ky_nang'
        if vi_tri_lam_viec is not None:
            vi_tri_lam_viec = 'kha_nang_lam_viec_sau'
        if thong_tin_chung is not None:
            thong_tin_chung = 'thong_tin_chung'

        if hoat_dong_doan is None and chi_tieu_cntt is None and khoi_xet_tuyen is None and co_so_dao_tao is None:
            if ky_thuat_may_tinh or cong_nghe_phan_mem or hung_yen_aptech or mang or thong_tin_chung or nhiem_vu_chuyen_mon or huong_nghien_cuu or doi_ngu_can_bo or kien_thuc_ky_nang or vi_tri_lam_viec:
                thong_tin = []
                if len(intent) >= 1:
                    print(3)
                    for i in range(len(intent)):
                        if intent[i] == 'khoa_cntt':
                            ky_thuat_may_tinh = regex_search(re_he_thong_nhung, sentence_activated[i])
                            cong_nghe_phan_mem = regex_search(re_lap_trinh_web, sentence_activated[i])
                            mang = regex_search(re_lap_trinh_mang, sentence_activated[i])
                            hung_yen_aptech = regex_search(re_hung_yen_aptech, sentence_activated[i])
                            thong_tin_chung = regex_search(re_thong_tin_chung, sentence_activated[i])
                            nhiem_vu_chuyen_mon = regex_search(re_nhiem_vu_chuyen_mon, sentence_activated[i])
                            huong_nghien_cuu = regex_search(re_huong_nghien_cuu, sentence_activated[i])
                            doi_ngu_can_bo = regex_search(re_doi_ngu_can_bo, sentence_activated[i])
                            kien_thuc_ky_nang = regex_search(re_kien_thuc_ky_nang, sentence_activated[i])
                            vi_tri_lam_viec = regex_search(re_vi_tri_lam_viec, sentence_activated[i])
                            if ky_thuat_may_tinh is not None:
                                ky_thuat_may_tinh = 'ky_thuat_may_tinh'
                            if cong_nghe_phan_mem is not None:
                                cong_nghe_phan_mem = 'cong_nghe_phan_mem'
                            if mang is not None:
                                mang = 'mang_may_tinh_va_truyen_thong'
                            if hung_yen_aptech is not None:
                                hung_yen_aptech = 'Hung_Yen_Aptech'
                            if nhiem_vu_chuyen_mon is not None:
                                nhiem_vu_chuyen_mon = 'nhiem_vu_chuyen_mon'
                            if huong_nghien_cuu is not None:
                                huong_nghien_cuu = 'cac_huong_nghien_cuu_chinh'
                            if doi_ngu_can_bo is not None:
                                doi_ngu_can_bo = 'doi_ngu_can_bo'
                            if kien_thuc_ky_nang is not None:
                                kien_thuc_ky_nang = 'kien_thuc_ky_nang'
                            if vi_tri_lam_viec is not None:
                                vi_tri_lam_viec = 'kha_nang_lam_viec_sau'
                            if thong_tin_chung is not None:
                                thong_tin_chung = 'thong_tin_chung'
                            x = checknotnull(ky_thuat_may_tinh, cong_nghe_phan_mem, mang, hung_yen_aptech)
                            y = checknotnull(nhiem_vu_chuyen_mon, huong_nghien_cuu, doi_ngu_can_bo, kien_thuc_ky_nang,
                                             vi_tri_lam_viec, thong_tin_chung)
                            if x is not None:
                                chuyen_nganh.append(x)
                            if y is not None:
                                thong_tin.append(y)
                    if len(chuyen_nganh) > 0 and len(thong_tin) > 0:
                        print(thong_tin[-1])
                        print(chuyen_nganh[-1])
                        result = give_answer(chuyen_nganh[-1], kichban=thong_tin[-1])
                        message = return_message(template_type='text', content=result)
                    if len(chuyen_nganh) == 0:
                        button1 = {'type': 'text', 'title': 'Công nghệ phần mềm'}
                        button2 = {'type': 'text', 'title': 'Kỹ thuật máy tính'}
                        button3 = {'type': 'text', 'title': 'Trung tâm Hưng Yên Aptech'}
                        button4 = {'type': 'text', 'title': 'Mạng máy tính và truyền thông'}
                        dict = add_button(button1, button2, button3, button4)
                        message = return_message(template_type='button',
                                                 content=[
                                                     'Hãy chọn một trong những thông tin sau để chúng tôi tư vấn:'],
                                                 template='template', dict=dict)
                    if len(thong_tin) == 0:
                        button1 = {'type': 'text', 'title': 'Thông tin chung'}
                        button2 = {'type': 'text', 'title': 'Nhiệm vụ chuyên môn'}
                        button3 = {'type': 'text', 'title': 'Các hướng nghiên cứu chính'}
                        button4 = {'type': 'text', 'title': 'Đội ngũ cán bộ'}
                        button5 = {'type': 'text', 'title': 'Kiến thức, kỹ năng'}
                        button6 = {'type': 'text', 'title': 'Vị trí làm việc sau khi tốt nghiệp'}
                        dict = add_button(button1, button2, button3, button4, button5, button6)
                        message = return_message(template_type='button',
                                                 content=['Chúng tôi có thể tư vấn giúp bạn về thông tin gì:'],
                                                 template='template', dict=dict)


            else:
                if ky_thuat_may_tinh is None and cong_nghe_phan_mem is None and hung_yen_aptech is None and mang is None:
                    button1 = {'type': 'text', 'title': 'Công nghệ phần mềm'}
                    button2 = {'type': 'text', 'title': 'Kỹ thuật máy tính'}
                    button3 = {'type': 'text', 'title': 'Trung tâm Hưng Yên Aptech'}
                    button4 = {'type': 'text', 'title': 'Mạng máy tính và truyền thông'}
                    dict = add_button(button1, button2, button3, button4)
                    message = return_message(template_type='button',
                                             content=['Hãy chọn một trong những thông tin sau để chúng tôi tư vấn:'],
                                             template='template', dict=dict)
                else:
                    if thong_tin_chung is None and nhiem_vu_chuyen_mon is None and huong_nghien_cuu is None and doi_ngu_can_bo is None and kien_thuc_ky_nang is None and vi_tri_lam_viec is None:
                        button1 = {'type': 'text', 'title': 'Thông tin chung'}
                        button2 = {'type': 'text', 'title': 'Nhiệm vụ chuyên môn'}
                        button3 = {'type': 'text', 'title': 'Các hướng nghiên cứu chính'}
                        button4 = {'type': 'text', 'title': 'Đội ngũ cán bộ'}
                        button5 = {'type': 'text', 'title': 'Kiến thức, kỹ năng'}
                        button6 = {'type': 'text', 'title': 'Vị trí làm việc sau khi tốt nghiệp'}
                        dict = add_button(button1, button2, button3, button4, button5, button6)
                        message = return_message(template_type='button',
                                                 content=['Chúng tôi có thể tư vấn giúp bạn về thông tin gì:'],
                                                 template='template', dict=dict)

        else:
            if hoat_dong_doan is not None:
                traloi = give_answer(kichban=hoat_dong_doan)
            if chi_tieu_cntt is not None or khoi_xet_tuyen is not None or co_so_dao_tao is not None:
                traloi = give_answer(chi_tieu_cntt, khoi_xet_tuyen, co_so_dao_tao, kichban='thong_tin_tuyen_sinh_cntt')
            message = return_message(template_type='text', content=traloi)

    else:
        result = give_answer(kichban=kichban)
        message = return_message(template_type='text', content=[ch(result)])

    return message


def return_message(template_type=None, content=None, template=None, dict=dict):
    message = {}
    attachment = {}
    payload = {}
    buttons = []
    data = {}

    message['message'] = data
    if template_type == 'text':
        data[template_type] = content
    elif template_type == 'button':
        message['message'] = attachment
        attachment['attachment'] = payload
        attachment['text'] = content
        payload['buttons'] = buttons
        for x in dict:
            buttons.append(x)

    return message


'''
def load_reply(name):
    sai = []
    file = "D:/TK 13.3/Doantotnghiep/Code/demo/file/reply/reply.xlsx"
    wb = op.load_workbook(file)
    x = wb.get_sheet_names()
    for i in x:
        if i == name:
            sheet = wb.get_sheet_by_name(name)
            for x in range(1, sheet.max_row + 1):
                if sheet.cell(row=x, column=1).value is not None:
                    sai.append(sheet.cell(row=x, column=1).value)
    return sai
'''


def add_button(*args):
    data = []
    for x in range(0, len(args)):
        data.append(args[x])
    return data


def handling(chuoi):
    # message = {}
    result = identify_intention(chuoi)
    for x in result:
        if x == 'status':
            if result[x] == 'Có':
                intent = result['name']

                '''data = load_reply(intent)
                if len(data) > 0:
                    message = return_message(template_type='text', content=ch(data))
                elif len(data) == 0:
                    button1 = {'type': 'text', 'title': 'ten'}
                    button2 = {'type': 'text', 'title': 'tuoi'}
                    button3 = {'type': 'text', 'title': 'hi'}
                    dict = add_button(button3, button1, button2)
                    message = return_message(template_type='button', content='Hãy chọn như cầu phù hợp với bạn',
                                             template='template', dict=dict)'''


            elif result[x] == 'Khong':
                intent = 'Sai'
                '''name = 'Sai'
                message = return_message(template_type='text', content=ch(load_reply(name)))
                '''
                wirte_file_excel.not_understand(chuoi)

    return intent
