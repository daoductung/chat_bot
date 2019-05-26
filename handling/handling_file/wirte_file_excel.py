#!/usr/bin/python3
# -*- coding: utf-8 -*-
import openpyxl as op


def not_understand(chuoi):
    mang = []
    wb = op.load_workbook('D:/TK 13.3/Doantotnghiep/Code/demo/file/not_understand/not_understand.xlsx')
    x = wb.get_sheet_names()

    my_wb = op.Workbook()
    my_sheet = my_wb.active
    sheet = wb.get_sheet_by_name(x[0])
    for x in range(1, sheet.max_row + 1):
        if sheet.cell(row=x, column=1).value is not None:
            mang.append(sheet.cell(row=x, column=1).value)
    mang.append(chuoi)

    for x in range(0, len(mang)):
            my_sheet.cell(row=x+1, column=1).value = mang[x]

    my_wb.save('D:/TK 13.3/Doantotnghiep/Code/demo/file/not_understand/not_understand.xlsx')
