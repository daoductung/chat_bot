#!/usr/bin/env python
# -*- coding: utf-8 -*-
import re
import openpyxl as op


def regex_search(regex, chuoi):
    t = re.search(regex, chuoi)

    if t is None:
        group = t.group(0)
        start = t.start(0)
        end = t.end(0)
    else:
        group = None
    return group


def give_answer(*args, kichban=str):
    cau_tra_loi = []
    file = "D:/TK 13.3/Doantotnghiep/Code/demo/file/reply/reply.xlsx"
    wb = op.load_workbook(file)
    x = wb.get_sheet_names()
    for i in x:
        if i == kichban:
            sheet = wb.get_sheet_by_name(kichban)
            if sheet:
                for x in range(1, sheet.max_row + 1):
                    for h in args:
                        if h != '':
                            if sheet.cell(row=x, column=2).value is not None and sheet.cell(row=x,
                                                                                            column=2).value == h:
                                cau_tra_loi.append(sheet.cell(row=x, column=1).value)
                            elif sheet.cell(row=x, column=2).value is None:
                                cau_tra_loi.append(sheet.cell(row=x, column=1).value)
    return cau_tra_loi


def return_reply(*args, body=str):
    intent = []  # Lưu tên kịch bản
    give_diem = 'https://diemthi.tuyensinh247.com/diem-chuan/dai-hoc-su-pham-ky-thuat-hung-yen-SKH.html?y='
    '''
        Phần regex tuyển sinh
    '''
    re_chi_tieu = r'(?<=)((chỉ\s+tiêu)|(bao nhiêu sinh viên)|(tuyển\s+chọn))'
    re_nhu_cau_diem = r'(?<=)(điểm)'
    re_diem = r'((điểm|diem)\s+\d{2})|(\d{2}\s+(điểm|diem))'
    re_nam = r'(?<=)(20\d+)'
    re_ho_so = r'(?<=)(hồ\s+sơ)'
    re_cach_nop_ho = r'(?<=)((nộp\s+hồ\s+sơ)|(nộp\s+vào\s+thời\s+gian)|(nộp))'
    re_phuong_thuc_tuyen_sinh = r'(?<=)((hình\s+thức)|(phương\s+thức)|(học\s+bạ)|(tuyển\s+sinh))'
    re_dieu_kien_xet_tuyen = r'(?<=)((điều kiện xét tuyển)|(dkxt))'

    '''
        Phần regex thông tin về trường
    '''

    re_lich_su_phat_trien = r'(?<=)((thành lập)|(từ năm))'
    re_vi_tri = r'(?<=)((nằm ở đâu)|(nằm chỗ nào)|(gần chỗ))'
    re_thong_tin = r'(?<=)((thông\s+tin)|(giới\s+thiệu\s+qua\s+về\s+trường)|(nổi\s+bật)|(thành\s+tích))'
    re_co_so_vat_chat = r'(?<=)((cơ\s+sở\s+vật\s+chất)|(đồ\s+dùng)|(cơ\s+sở)|(sân)|(thể\s+thao))'
    re_hoc_phi = r'(?<=)(học\s+phí)'
    re_co_hoi_viec_lam = r'(?<=)((việc\s+làm)|(cơ\s+hội\s+việc\s+làm))'
    re_hoat_dong_doan_the = r'(?<=)((hoạt\s+động)|đoàn)'

    '''
        Phần regex thông tin ngành học
    '''
    re_he_thong_nhung = r'(?<=)(nhúng|hệ thống nhúng|làm mạch|điều khiển|iot|smart\s+home|công nghệ máy tính|kỹ thuật máy tính)'
    re_lap_trinh_web = r'(?<=)(công nghệ phần mềm|phần mềm|web|website)'
    re_lap_trinh_mang = r'(?<=)(mạng)'
    re_lap_trinh_mobile = r'(?<=)(di động|mobile|android|ios|ứng dụng|app)'
    re_kiem_thu = r'(?<=)(kiểm thử phần mềm|tester|test)'

    '''
    '''
    thong_tin_truong = []  # Nhánh thông tin trường
    lich_su_pt = []  # Lịch sử phát triển
    vi_tri = []  # Vị trí
    thong_tin_noi_bat = []  # Thông tin nổi bật của trường
    co_so_vat_chat = []  # Cơ sở vật chất của trường
    hoc_phi = []  # Học phí của trường
    co_hoi_viec_lam = []  # Cơ hội việc làm sau khi ra trường
    hoat_dong = []  # Hoạt động của trường

    '''
    '''
    thong_tin_tuyensinh = []  # Nhánh thông tin tuyển sinh
    diem_chuan = []  # Điểm chuẩn của năm
    diem_chuan_nam = []  # Lưu số năm
    phuong_thuc_tuyen_sinh = []  # Phương thức tuyển sinh
    ho_so = []  # Hồ sơ tuyển sinh
    cach_nop_ho_so = []  # Cách nộp hồ sơ tuyển sinh
    chi_tieu_tuyen_sinh = []  # Chỉ tiêu tuyển sinh của nhà trường
    '''
    '''
    thong_tin_nganhhoc = []  # Nhánh ngành học
    ten_nganh = []  # Tên ngành
    gioi_thieu_nganh = []  # Giới thiệu về ngành
    chi_tieu_tuyen_sinh_khoa = []  # Chỉ tiêu của khoa
    hoat_dong_khoa = []  # Hoạt động ngoại khóa của khoa
    co_hoi_viec_lam_khoa = []  # Cơ hội việc làm sau khi hoạt đông

    gt_thong_tin_chung = []  # Giới thiệu thông chung về ngành học
    gt_nhiem_vu_chuyen_mon = []  # Giới thiệu nhiệm vụ chuyên môn
    gt_huong_nghien_cuu = []  # Giới thiệu hướng nghiên cứu
    gt_doi_ngu_gv = []  # Giới thiệu đội ngũ giáo viên

    print(args)
    it = len(args[0])
    print(it)
    if it >= 1:
        if args[0][it - 2] == 'thong_tin_ve_truong':
            lich_su_pt = regex_search(re_lich_su_phat_trien, body)
            vi_tri = regex_search(re_vi_tri, body)
            thong_tin_noi_bat = regex_search(re_thong_tin, body)
            co_so_vat_chat = regex_search(re_co_so_vat_chat, body)
            hoc_phi = regex_search(re_hoc_phi, body)
            co_hoi_viec_lam = regex_search(re_co_hoi_viec_lam, body)
            hoat_dong = regex_search(re_hoat_dong_doan_the, body)
            if lich_su_pt is not None:
                lich_su_pt = 'lich_su_phat_trien'
            else:
                lich_su_pt = ''
            if vi_tri is not None:
                vi_tri = 'vi_tri'
            else:
                vi_tri = ''
            if thong_tin_noi_bat is not None:
                thong_tin_noi_bat = 'giai_thuong'
            else:
                thong_tin_noi_bat = ''
            if co_so_vat_chat is not None:
                co_so_vat_chat = 'co_so_vat_chat'
            else:
                co_so_vat_chat = ''
            if hoc_phi is not None:
                hoc_phi = 'hoc_phi'
            else:
                hoc_phi = ''
            if co_hoi_viec_lam is not None:
                co_hoi_viec_lam = 'co_hoi_viec_lam'
            else:
                co_hoi_viec_lam = ''
            if hoat_dong is not None:
                hoat_dong = 'hoat_dong_truong'
            else:
                hoat_dong = ''
            give_answer(lich_su_pt, vi_tri, thong_tin_noi_bat, co_so_vat_chat, hoc_phi, co_hoi_viec_lam, hoat_dong,
                        kichban='thong_tin_ve_truong')
        elif args[0][it - 2] == 'tuyen_sinh':
            chi_tieu_tuyen_sinh = regex_search(re_chi_tieu, body)
            nhu_cau_diem = regex_search(re_nhu_cau_diem, body)
            ho_so = regex_search(re_ho_so, body)
            cach_nop_ho_so = regex_search(re_cach_nop_ho, body)
            phuong_thuc_tuyen_sinh = regex_search(re_phuong_thuc_tuyen_sinh, body)
            dieu_kien_xet_tuyen = regex_search(re_dieu_kien_xet_tuyen, body)
            if chi_tieu_tuyen_sinh is not None:
                chi_tieu_tuyen_sinh = 'chi_tieu_tuyen_sinh'
            else:
                chi_tieu_tuyen_sinh = ''
            if nhu_cau_diem is not None:
                nhu_cau_diem = 'diem_tuyen_sinh'
                nam = regex_search(re_nam, body)
                diem = regex_search(re_diem, body)

            else:
                diem = ''
            if ho_so is not None:
                ho_so = 'ho_so_tuyen_sinh'
            else:
                ho_so = ''
            if cach_nop_ho_so is not None:
                cach_nop_ho_so = 'cach_nop_ho_so'
            else:
                cach_nop_ho_so = ''
            if phuong_thuc_tuyen_sinh is not None:
                phuong_thuc_tuyen_sinh = 'phuong_thuc_tuyen_sinh'
            else:
                phuong_thuc_tuyen_sinh = ''
            if dieu_kien_xet_tuyen is not None:
                dieu_kien_xet_tuyen = 'dieu_kien_xet_tuyen'
            else:
                dieu_kien_xet_tuyen = ''

        elif args[0][it - 2] == 'khoa_cntt':
            chuyen_nganh = []
            he_thong_nhung = regex_search(re_he_thong_nhung, body)
            web = regex_search(re_lap_trinh_web, body)
            kiem_thu = regex_search(re_kiem_thu, body)
            mang = regex_search(re_lap_trinh_mang, body)
            mobile = regex_search(re_lap_trinh_mobile, body)

    return 1
